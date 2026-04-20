"""
Tabulaire export van audit-bevindingen — CSV + Excel (.xlsx).

Naast het Markdown-rapport (local_report.py) levert deze module een platte
tabelvorm geschikt voor filtering, Sheets, en audit-trail.

Bundeling (route A — heuristisch):
  Elke bevinding krijgt een thema-label op basis van keyword-matching in
  `beschrijving`. First-match-wins. Thema's staan in THEMA_REGELS.

Gebruik:
  # standalone, leest bevindingen uit output/audit.db
  python3 -m audit.tabular_report --norm beide
  python3 -m audit.tabular_report --norm beide --scherpte 0.5

  # programmatisch (aanroep vanuit pipeline.py)
  from audit.tabular_report import schrijf_csv, schrijf_excel
  schrijf_csv(bevindingen, norm="beide")
  schrijf_excel(bevindingen, norm="beide")
"""

import argparse
import csv
import logging
import os
import re
import sqlite3
from collections import Counter, defaultdict
from datetime import date

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from audit.normteksten import NORMTEKSTEN_9001, NORMTEKSTEN_27001

logger = logging.getLogger(__name__)

DEFAULT_OUTPUT_DIR = os.path.join(
    os.path.dirname(os.path.dirname(__file__)), "output", "audit_reports"
)
DEFAULT_DB_PATH = os.path.join(
    os.path.dirname(os.path.dirname(__file__)), "output", "audit.db"
)

VOLGORDE_CLASSIFICATIE = {"NC": 0, "OFI": 1, "positief": 2}

# Vaste taxonomie — gedeeld met thema_classifier.py (route B).
# "Overig" is de fallback-categorie.
THEMA_LIJST: list[str] = [
    "Offboarding & activa-retournering",
    "Informatieclassificatie",
    "Cryptografie & encryptie",
    "Back-up & continuïteit",
    "Screening & HR",
    "Fysieke beveiliging",
    "Leveranciersbeheer",
    "Memo & afwijkingsregistratie",
    "Template zonder toepassing",
    "Verificatie van doeltreffendheid",
    "Klanttevredenheid",
    "Documentatie-actualiteit",
    "Rollen & verantwoordelijkheden",
    "Beleid",
    "Risicomanagement",
    "Auditprogramma",
    "Directiebeoordeling",
    "Training & competenties",
    "Toegangsbeheer",
    "Logging & monitoring",
    "Incident response",
    "Wettelijke & contractuele eisen",
    "Privacy & persoonsgegevens",
    "Context-analyse & belanghebbenden",
    "Overig",
]

# Keyword-gedreven thema-regels (route A). First-match-wins — specifieker eerst.
# Elke bevinding krijgt precies één thema; geen match → "Overig".
THEMA_REGELS: list[tuple[str, list[str]]] = [
    ("Offboarding & activa-retournering",
     ["retournering", "offboarding", "beëindiging dienst", "teruggave activa"]),
    ("Informatieclassificatie",
     ["informatieclassificatie", "classificatie van informatie", "label"]),
    ("Cryptografie & encryptie",
     ["cryptograf", "encryptie", "versleuteling"]),
    ("Back-up & continuïteit",
     ["back-up", "backup", "continuïteit", "disaster recovery", "bcm"]),
    ("Screening & HR",
     ["screening", "achtergrondcontrole", "arbeidsvoorwaarden"]),
    ("Fysieke beveiliging",
     ["perimeter", "fysieke beveilig", "kantoorbeveilig"]),
    ("Leveranciersbeheer",
     ["leverancier", "supplier", "derde partij", "uitbesteding"]),
    ("Memo & afwijkingsregistratie",
     ["memo afwijk", "memo nc", "afwijkingsprocedure", "afwijkingsmemo", "nc-memo"]),
    ("Template zonder toepassing",
     ["template", "ongevuld", "niet ingevuld", "geen concrete", "lege procedure"]),
    ("Verificatie van doeltreffendheid",
     ["verificatie", "effectiviteit", "doeltreffendheid", "follow-up"]),
    ("Klanttevredenheid",
     ["klanttevredenheid", "nps ", "klantbeoordeling", "klantreview"]),
    ("Documentatie-actualiteit",
     ["verouderd", "niet meer actueel", "ouder dan 2 jaar", "> 2 jaar",
      "niet geactualiseerd", "reeds geactualiseerd", "herziening nodig"]),
    ("Rollen & verantwoordelijkheden",
     ["rollen", "verantwoordelijkheden", "bevoegdheden"]),
    ("Beleid",
     ["kwaliteitsbeleid", "informatiebeveiligingsbeleid", "beleidsdocument", "beleidsverklaring"]),
    ("Risicomanagement",
     ["risicobeoordeling", "risicoanalyse", "risicoregister"]),
    ("Wettelijke & contractuele eisen",
     ["wet en regelgeving", "wettelijke eis", "contractuele eis", "compliance",
      "regelgeving", "iso-704"]),
    ("Privacy & persoonsgegevens",
     ["privacy", "avg", "gdpr", "persoonsgegevens", "pii"]),
    ("Context-analyse & belanghebbenden",
     ["contextanalyse", "context analyse", "belanghebbenden", "stakeholder",
      "swot", "interne en externe context"]),
    ("Auditprogramma",
     ["interne audit", "auditprogramma", "auditplan"]),
    ("Directiebeoordeling",
     ["directiebeoordeling", "management review"]),
    ("Training & competenties",
     ["training", "competenties", "opleiding"]),
    ("Toegangsbeheer",
     ["toegangsbeheer", "toegangsrechten", "autorisatie"]),
    ("Logging & monitoring",
     ["logging", "monitoring", "audit-trail", "audit trail"]),
    ("Incident response",
     ["incident response", "incidentmanagement", "incidentprocedure"]),
]


def _bepaal_norm_voor_clausule(clausule_id: str) -> str:
    """
    Return "9001", "27001" of "beide" op basis van de normteksten-lookup.
    Alleen-in-9001 → 9001. Alleen-in-27001 → 27001. Beide → beide.
    """
    in_9001 = clausule_id in NORMTEKSTEN_9001
    in_27001 = clausule_id in NORMTEKSTEN_27001
    if in_9001 and in_27001:
        return "beide"
    if in_9001:
        return "9001"
    if in_27001:
        return "27001"
    return "onbekend"


def bepaal_thema(bevinding: dict) -> str:
    """Keyword-match over beschrijving + onderbouwing. First-match-wins."""
    tekst = " ".join([
        bevinding.get("beschrijving") or "",
        bevinding.get("onderbouwing") or "",
        bevinding.get("document_naam") or "",
    ]).lower()

    if not tekst.strip():
        return "Overig"

    for thema, keywords in THEMA_REGELS:
        for kw in keywords:
            if kw.lower() in tekst:
                return thema
    return "Overig"


def _doc_url(bevinding: dict) -> str:
    doc_id = bevinding.get("doc_id", "")
    if not doc_id:
        return ""
    if bevinding.get("herkomst") == "Miro":
        board_id = os.environ.get("MIRO_BOARD_ID", "")
        if board_id:
            return f"https://miro.com/app/board/{board_id}/?moveToWidget={doc_id}"
        return ""
    return f"https://drive.google.com/file/d/{doc_id}/view"


def _verrijk(
    bevindingen: list[dict],
    llm_themas: dict[str, str] | None = None,
) -> list[dict]:
    """
    Voeg afgeleide kolommen toe: norm, thema, thema_bron, doc_url.

    Als llm_themas is meegegeven, wordt voor matching IDs het LLM-thema gebruikt
    (met `thema_bron="llm"`). Anders heuristiek (`thema_bron="heuristisch"`).
    """
    verrijkt = []
    for i, bev in enumerate(bevindingen):
        clausule = bev.get("clausule") or bev.get("clausule_id", "")
        bev_id = str(bev.get("id") or bev.get("_bev_id") or i)
        row = dict(bev)
        row["_bev_id"] = bev_id
        row["clausule"] = clausule
        row["norm"] = _bepaal_norm_voor_clausule(clausule)
        if llm_themas and bev_id in llm_themas:
            row["thema"] = llm_themas[bev_id]
            row["thema_bron"] = "llm"
        else:
            row["thema"] = bepaal_thema(bev)
            row["thema_bron"] = "heuristisch"
        row["doc_url"] = _doc_url(bev)
        verrijkt.append(row)
    return verrijkt


def _sorteer(bevindingen: list[dict]) -> list[dict]:
    """Sorteer: clausule numeriek, dan classificatie (NC → OFI → positief), dan thema."""
    def clausule_sleutel(c: str) -> tuple:
        # "10.2" → (10, 2); "5.11" → (5, 11)
        delen = re.findall(r"\d+", c)
        return tuple(int(d) for d in delen) if delen else (9999,)

    return sorted(
        bevindingen,
        key=lambda b: (
            clausule_sleutel(b.get("clausule", "")),
            VOLGORDE_CLASSIFICATIE.get(b.get("classificatie", ""), 9),
            b.get("thema", ""),
        ),
    )


def _bestandsnaam(prefix: str, norm: str, scherpte: float, extensie: str) -> str:
    norm_deel = norm.replace(" ", "").replace("+", "-")
    scherpte_label = f"_s{str(scherpte).replace('.', '')}" if scherpte != 1.0 else ""
    return f"{prefix}_{norm_deel}_{date.today()}{scherpte_label}.{extensie}"


CSV_KOLOMMEN = [
    "norm",
    "clausule",
    "clausule_titel",
    "classificatie",
    "thema",
    "thema_bron",
    "document_naam",
    "herkomst",
    "beschrijving",
    "onderbouwing",
    "doc_id",
    "doc_url",
    "classified_at",
]


def schrijf_csv(
    bevindingen: list[dict],
    norm: str,
    scherpte: float = 1.0,
    output_dir: str | None = None,
    llm_themas: dict[str, str] | None = None,
) -> str:
    """Schrijf bevindingen als platte CSV. Retourneert het pad."""
    output_dir = output_dir or os.environ.get("LOCAL_REPORT_DIR", DEFAULT_OUTPUT_DIR)
    os.makedirs(output_dir, exist_ok=True)

    pad = os.path.join(output_dir, _bestandsnaam("Bevindingen", norm, scherpte, "csv"))
    rijen = _sorteer(_verrijk(bevindingen, llm_themas=llm_themas))

    with open(pad, "w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=CSV_KOLOMMEN, extrasaction="ignore")
        writer.writeheader()
        for r in rijen:
            writer.writerow({k: r.get(k, "") for k in CSV_KOLOMMEN})

    logger.info("CSV geschreven: %s (%d rijen)", pad, len(rijen))
    return pad


CLASSIFICATIE_KLEUREN = {
    "NC": "FFC7CE",        # rood
    "OFI": "FFEB9C",       # oranje
    "positief": "C6EFCE",  # groen
}


def schrijf_excel(
    bevindingen: list[dict],
    norm: str,
    scherpte: float = 1.0,
    output_dir: str | None = None,
    llm_themas: dict[str, str] | None = None,
) -> str:
    """
    Schrijf bevindingen als .xlsx met drie tabs:
      1. Samenvatting    — aantal per (norm, thema, classificatie)
      2. Bevindingen     — platte tabel met alle kolommen, gekleurd per NC/OFI/pos
      3. Per clausule    — aantal NC/OFI/positief per clausule
    """
    output_dir = output_dir or os.environ.get("LOCAL_REPORT_DIR", DEFAULT_OUTPUT_DIR)
    os.makedirs(output_dir, exist_ok=True)

    pad = os.path.join(output_dir, _bestandsnaam("Bevindingen", norm, scherpte, "xlsx"))
    rijen = _sorteer(_verrijk(bevindingen, llm_themas=llm_themas))

    wb = Workbook()
    _tab_samenvatting(wb.active, rijen)
    _tab_bevindingen(wb.create_sheet("Bevindingen"), rijen)
    _tab_per_clausule(wb.create_sheet("Per clausule"), rijen)

    wb.save(pad)
    logger.info("Excel geschreven: %s (%d rijen)", pad, len(rijen))
    return pad


def _tab_samenvatting(ws, rijen: list[dict]) -> None:
    ws.title = "Samenvatting"
    header_font = Font(bold=True)

    # Totalen per classificatie
    ws.append(["Totalen"])
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws.append(["Classificatie", "Aantal"])
    for c in ws[2]:
        c.font = header_font
    teller_cls = Counter(r["classificatie"] for r in rijen)
    for cls in ("NC", "OFI", "positief"):
        ws.append([cls, teller_cls.get(cls, 0)])
    ws.append(["Totaal", len(rijen)])
    ws[ws.max_row][0].font = header_font
    ws[ws.max_row][1].font = header_font
    ws.append([])

    # Per thema
    ws.append(["Bundeling per thema"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)
    ws.append(["Thema", "NC", "OFI", "Positief", "Totaal"])
    for c in ws[ws.max_row]:
        c.font = header_font
    teller_thema: dict[str, Counter] = defaultdict(Counter)
    for r in rijen:
        teller_thema[r["thema"]][r["classificatie"]] += 1
    # Sorteer op totaal afnemend, "Overig" laatst
    def thema_sleutel(item):
        thema, teller = item
        totaal = sum(teller.values())
        return (thema == "Overig", -totaal, thema)
    for thema, teller in sorted(teller_thema.items(), key=thema_sleutel):
        ws.append([
            thema,
            teller.get("NC", 0),
            teller.get("OFI", 0),
            teller.get("positief", 0),
            sum(teller.values()),
        ])
    ws.append([])

    # Per norm
    ws.append(["Per norm"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)
    ws.append(["Norm", "NC", "OFI", "Positief", "Totaal"])
    for c in ws[ws.max_row]:
        c.font = header_font
    teller_norm: dict[str, Counter] = defaultdict(Counter)
    for r in rijen:
        teller_norm[r["norm"]][r["classificatie"]] += 1
    for norm_key in ("9001", "27001", "beide", "onbekend"):
        if norm_key in teller_norm:
            t = teller_norm[norm_key]
            ws.append([
                norm_key, t.get("NC", 0), t.get("OFI", 0), t.get("positief", 0), sum(t.values())
            ])

    _auto_width(ws)


def _tab_bevindingen(ws, rijen: list[dict]) -> None:
    # Header
    ws.append(CSV_KOLOMMEN)
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="D9D9D9")

    for r in rijen:
        ws.append([r.get(k, "") for k in CSV_KOLOMMEN])

    # Conditional formatting op kolom "classificatie"
    cls_col_idx = CSV_KOLOMMEN.index("classificatie") + 1
    cls_letter = get_column_letter(cls_col_idx)
    rng = f"{cls_letter}2:{cls_letter}{ws.max_row}"
    for cls, kleur in CLASSIFICATIE_KLEUREN.items():
        ws.conditional_formatting.add(
            rng,
            CellIsRule(operator="equal", formula=[f'"{cls}"'], fill=PatternFill("solid", fgColor=kleur)),
        )

    # Wrap beschrijving + onderbouwing
    for kolom_naam in ("beschrijving", "onderbouwing"):
        idx = CSV_KOLOMMEN.index(kolom_naam) + 1
        for row_cells in ws.iter_rows(min_row=2, min_col=idx, max_col=idx):
            for c in row_cells:
                c.alignment = Alignment(wrap_text=True, vertical="top")

    # Freeze header + eerste 2 kolommen
    ws.freeze_panes = "C2"

    _auto_width(ws, max_breedte=60)


def _tab_per_clausule(ws, rijen: list[dict]) -> None:
    ws.append(["Clausule", "Norm", "Titel", "NC", "OFI", "Positief", "Totaal"])
    for c in ws[1]:
        c.font = Font(bold=True)

    teller: dict[tuple[str, str], Counter] = defaultdict(Counter)
    titels: dict[str, str] = {}
    for r in rijen:
        sleutel = (r["clausule"], r["norm"])
        teller[sleutel][r["classificatie"]] += 1
        if r.get("clausule_titel"):
            titels[r["clausule"]] = r["clausule_titel"]

    def clausule_sleutel(c: str) -> tuple:
        delen = re.findall(r"\d+", c)
        return tuple(int(d) for d in delen) if delen else (9999,)

    for (clausule, norm), t in sorted(teller.items(), key=lambda kv: clausule_sleutel(kv[0][0])):
        ws.append([
            clausule,
            norm,
            titels.get(clausule, ""),
            t.get("NC", 0),
            t.get("OFI", 0),
            t.get("positief", 0),
            sum(t.values()),
        ])

    ws.freeze_panes = "A2"
    _auto_width(ws, max_breedte=50)


def _auto_width(ws, max_breedte: int = 40) -> None:
    for kolom_cellen in ws.columns:
        max_lengte = 0
        letter = None
        for cel in kolom_cellen:
            if letter is None:
                letter = cel.column_letter
            waarde = str(cel.value) if cel.value is not None else ""
            # eerste regel telt (multiline strings niet meerekenen voor breedte)
            eerste_regel = waarde.split("\n", 1)[0]
            if len(eerste_regel) > max_lengte:
                max_lengte = len(eerste_regel)
        if letter:
            ws.column_dimensions[letter].width = min(max_lengte + 2, max_breedte)


def lees_uit_db(db_pad: str | None = None, norm: str | None = None) -> list[dict]:
    """
    Lees bevindingen uit audit.db en normaliseer naar het in-memory schema
    (zelfde keys als local_report.py gebruikt).
    """
    db_pad = db_pad or os.environ.get("AUDIT_DB_PATH", DEFAULT_DB_PATH)
    if not os.path.exists(db_pad):
        raise FileNotFoundError(f"audit.db niet gevonden: {db_pad}")

    conn = sqlite3.connect(db_pad)
    conn.row_factory = sqlite3.Row
    try:
        query = "SELECT * FROM bevindingen"
        params: tuple = ()
        if norm and norm != "beide":
            query += " WHERE norm = ?"
            params = (norm,)
        rows = conn.execute(query, params).fetchall()
    finally:
        conn.close()

    bevindingen = []
    for r in rows:
        bevindingen.append({
            "id": r["id"],
            "clausule": r["clausule_id"],
            "clausule_id": r["clausule_id"],
            "clausule_titel": "",  # niet in DB; blijft leeg
            "norm": r["norm"],
            "classificatie": r["classificatie"],
            "beschrijving": r["beschrijving"] or "",
            "onderbouwing": r["onderbouwing"] or "",
            "document_naam": r["document_naam"] or "",
            "herkomst": r["herkomst"],
            "doc_id": r["doc_id"],
            "classified_at": r["classified_at"],
        })
    return bevindingen


def main():
    parser = argparse.ArgumentParser(description="Tabulaire export van audit-bevindingen")
    parser.add_argument("--norm", choices=["9001", "27001", "beide"], default="beide")
    parser.add_argument("--scherpte", type=float, default=1.0)
    parser.add_argument("--db", default=None, help="Pad naar audit.db (default: output/audit.db)")
    parser.add_argument("--output-dir", default=None)
    parser.add_argument("--alleen-csv", action="store_true")
    parser.add_argument("--alleen-excel", action="store_true")
    parser.add_argument(
        "--thema-llm",
        action="store_true",
        help="Gebruik LLM voor thema-toekenning op findings waar heuristiek 'Overig' geeft",
    )
    args = parser.parse_args()

    logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")

    bevindingen = lees_uit_db(args.db, args.norm)
    logger.info("%d bevindingen geladen uit DB", len(bevindingen))

    llm_themas: dict[str, str] = {}
    if args.thema_llm:
        from audit.thema_classifier import verfijn_overig
        llm_themas = verfijn_overig(bevindingen)

    if not args.alleen_excel:
        schrijf_csv(bevindingen, args.norm, args.scherpte, args.output_dir,
                    llm_themas=llm_themas)
    if not args.alleen_csv:
        schrijf_excel(bevindingen, args.norm, args.scherpte, args.output_dir,
                      llm_themas=llm_themas)


if __name__ == "__main__":
    main()
