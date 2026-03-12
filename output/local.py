#!/usr/bin/env python3
"""Write merged ArgoCD rows to a local .xlsx file via openpyxl."""

import json
import os
import sys

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from upsert import OWNED_COLUMNS, upsert

OUTPUT_PATH = os.environ.get("LOCAL_OUTPUT_PATH", "argocd_apps.xlsx")


def read_existing() -> tuple[list[str], list[dict]]:
    if not os.path.exists(OUTPUT_PATH):
        return [], []

    wb = openpyxl.load_workbook(OUTPUT_PATH)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []

    headers = [str(h) for h in rows[0]]
    data = [dict(zip(headers, row)) for row in rows[1:]]
    return headers, data


def write_sheet(headers: list[str], rows: list[dict]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append([("" if row.get(h) is None else row.get(h)) for h in headers])
    wb.save(OUTPUT_PATH)


def main(new_rows_path: str) -> None:
    with open(new_rows_path) as f:
        new_rows = json.load(f)

    existing_headers, current_rows = read_existing()

    # Owned columns first; preserve any extra columns users added manually.
    all_headers = list(OWNED_COLUMNS)
    for h in existing_headers:
        if h not in all_headers:
            all_headers.append(h)

    merged = upsert(current_rows, new_rows)
    write_sheet(all_headers, merged)
    print(f"[local] Written {len(merged)} rows to {OUTPUT_PATH}")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: output/local.py <new_rows.json>", file=sys.stderr)
        sys.exit(1)
    main(sys.argv[1])
